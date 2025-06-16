"""
Data Export Functionality for EOI System - ISG 2026 Volunteer Management

This module provides comprehensive data export capabilities for staff to export
EOI submissions in various formats (CSV, Excel, PDF) with filtering and customization options.
"""

import csv
import json
import io
from datetime import datetime, date
from django.http import HttpResponse, JsonResponse
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.utils.translation import gettext_lazy as _
from django.db.models import Q, Count, Avg
from django.template.loader import render_to_string
from django.conf import settings
import logging

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

from .eoi_models import EOISubmission, EOIProfileInformation, EOIRecruitmentPreferences, EOIGamesInformation
from .permissions import IsStaffOrVMTOrCVT
from common.audit import log_audit_event

logger = logging.getLogger(__name__)


class EOIDataExporter:
    """
    Main class for exporting EOI data in various formats
    """
    
    def __init__(self, user, export_format='csv'):
        self.user = user
        self.export_format = export_format.lower()
        self.timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Verify user permissions
        if not self._check_permissions():
            raise PermissionDenied("User does not have permission to export data")
    
    def _check_permissions(self):
        """Check if user has permission to export data"""
        permission_checker = IsStaffOrVMTOrCVT()
        # Create a mock request object for permission checking
        class MockRequest:
            def __init__(self, user):
                self.user = user
        
        mock_request = MockRequest(self.user)
        return permission_checker.has_permission(mock_request, None)
    
    def get_base_queryset(self, filters=None):
        """Get base queryset with optional filters"""
        queryset = EOISubmission.objects.select_related(
            'profile_information',
            'recruitment_preferences', 
            'games_information',
            'user'
        ).prefetch_related(
            'profile_information',
            'recruitment_preferences',
            'games_information'
        )
        
        if filters:
            queryset = self._apply_filters(queryset, filters)
        
        return queryset.order_by('-created_at')
    
    def _apply_filters(self, queryset, filters):
        """Apply filters to queryset"""
        # Status filter
        if filters.get('status'):
            queryset = queryset.filter(status__in=filters['status'])
        
        # Volunteer type filter
        if filters.get('volunteer_type'):
            queryset = queryset.filter(volunteer_type__in=filters['volunteer_type'])
        
        # Date range filter
        if filters.get('date_from'):
            queryset = queryset.filter(created_at__gte=filters['date_from'])
        if filters.get('date_to'):
            queryset = queryset.filter(created_at__lte=filters['date_to'])
        
        # Email confirmation filter
        if filters.get('email_confirmed') is not None:
            queryset = queryset.filter(email_confirmed=filters['email_confirmed'])
        
        # Completion status filter
        if filters.get('completion_status'):
            if filters['completion_status'] == 'complete':
                queryset = queryset.filter(
                    profile_section_complete=True,
                    recruitment_section_complete=True,
                    games_section_complete=True
                )
            elif filters['completion_status'] == 'incomplete':
                queryset = queryset.filter(
                    Q(profile_section_complete=False) |
                    Q(recruitment_section_complete=False) |
                    Q(games_section_complete=False)
                )
        
        # Search filter
        if filters.get('search'):
            search_term = filters['search']
            queryset = queryset.filter(
                Q(profile_information__first_name__icontains=search_term) |
                Q(profile_information__last_name__icontains=search_term) |
                Q(profile_information__email__icontains=search_term) |
                Q(profile_information__phone_number__icontains=search_term)
            )
        
        return queryset
    
    def get_field_definitions(self):
        """Define available fields for export"""
        return {
            # Basic submission info
            'submission_id': {'label': 'Submission ID', 'category': 'Basic'},
            'volunteer_type': {'label': 'Volunteer Type', 'category': 'Basic'},
            'status': {'label': 'Status', 'category': 'Basic'},
            'created_at': {'label': 'Created Date', 'category': 'Basic'},
            'submitted_at': {'label': 'Submitted Date', 'category': 'Basic'},
            'email_confirmed': {'label': 'Email Confirmed', 'category': 'Basic'},
            
            # Profile information
            'first_name': {'label': 'First Name', 'category': 'Profile'},
            'last_name': {'label': 'Last Name', 'category': 'Profile'},
            'preferred_name': {'label': 'Preferred Name', 'category': 'Profile'},
            'email': {'label': 'Email', 'category': 'Profile'},
            'phone_number': {'label': 'Phone Number', 'category': 'Profile'},
            'date_of_birth': {'label': 'Date of Birth', 'category': 'Profile'},
            'age': {'label': 'Age', 'category': 'Profile'},
            'gender': {'label': 'Gender', 'category': 'Profile'},
            'address_full': {'label': 'Full Address', 'category': 'Profile'},
            'city': {'label': 'City', 'category': 'Profile'},
            'country': {'label': 'Country', 'category': 'Profile'},
            'emergency_contact_name': {'label': 'Emergency Contact Name', 'category': 'Profile'},
            'emergency_contact_phone': {'label': 'Emergency Contact Phone', 'category': 'Profile'},
            'education_level': {'label': 'Education Level', 'category': 'Profile'},
            'employment_status': {'label': 'Employment Status', 'category': 'Profile'},
            'occupation': {'label': 'Occupation', 'category': 'Profile'},
            'languages_spoken': {'label': 'Languages Spoken', 'category': 'Profile'},
            'nationality': {'label': 'Nationality', 'category': 'Profile'},
            
            # Recruitment preferences
            'volunteer_experience_level': {'label': 'Experience Level', 'category': 'Recruitment'},
            'previous_events': {'label': 'Previous Events', 'category': 'Recruitment'},
            'special_skills': {'label': 'Special Skills', 'category': 'Recruitment'},
            'motivation': {'label': 'Motivation', 'category': 'Recruitment'},
            'preferred_sports': {'label': 'Preferred Sports', 'category': 'Recruitment'},
            'preferred_venues': {'label': 'Preferred Venues', 'category': 'Recruitment'},
            'preferred_roles': {'label': 'Preferred Roles', 'category': 'Recruitment'},
            'availability_level': {'label': 'Availability Level', 'category': 'Recruitment'},
            'max_hours_per_day': {'label': 'Max Hours Per Day', 'category': 'Recruitment'},
            'has_own_transport': {'label': 'Has Own Transport', 'category': 'Recruitment'},
            'transport_method': {'label': 'Transport Method', 'category': 'Recruitment'},
            
            # Games information
            't_shirt_size': {'label': 'T-Shirt Size', 'category': 'Games'},
            'dietary_requirements': {'label': 'Dietary Requirements', 'category': 'Games'},
            'has_food_allergies': {'label': 'Has Food Allergies', 'category': 'Games'},
            'food_allergy_details': {'label': 'Food Allergy Details', 'category': 'Games'},
            'requires_accommodation': {'label': 'Requires Accommodation', 'category': 'Games'},
            'preferred_shifts': {'label': 'Preferred Shifts', 'category': 'Games'},
            'is_part_of_group': {'label': 'Part of Group', 'category': 'Games'},
            'group_name': {'label': 'Group Name', 'category': 'Games'},
            'how_did_you_hear': {'label': 'How Did You Hear', 'category': 'Games'},
        }
    
    def extract_data(self, submission, fields):
        """Extract data from submission based on requested fields"""
        data = {}
        
        for field in fields:
            try:
                if field == 'submission_id':
                    data[field] = str(submission.id)
                elif field == 'volunteer_type':
                    data[field] = submission.get_volunteer_type_display()
                elif field == 'status':
                    data[field] = submission.get_status_display()
                elif field == 'created_at':
                    data[field] = submission.created_at.strftime('%Y-%m-%d %H:%M:%S')
                elif field == 'submitted_at':
                    data[field] = submission.submitted_at.strftime('%Y-%m-%d %H:%M:%S') if submission.submitted_at else ''
                elif field == 'email_confirmed':
                    data[field] = 'Yes' if submission.email_confirmed else 'No'
                
                # Profile information fields
                elif hasattr(submission, 'profile_information') and submission.profile_information:
                    profile = submission.profile_information
                    if field == 'first_name':
                        data[field] = profile.first_name or ''
                    elif field == 'last_name':
                        data[field] = profile.last_name or ''
                    elif field == 'preferred_name':
                        data[field] = profile.preferred_name or ''
                    elif field == 'email':
                        data[field] = profile.email or ''
                    elif field == 'phone_number':
                        data[field] = profile.phone_number or ''
                    elif field == 'date_of_birth':
                        data[field] = profile.date_of_birth.strftime('%Y-%m-%d') if profile.date_of_birth else ''
                    elif field == 'age':
                        data[field] = profile.age if hasattr(profile, 'age') else ''
                    elif field == 'gender':
                        data[field] = profile.get_gender_display() if profile.gender else ''
                    elif field == 'address_full':
                        address_parts = [
                            profile.address_line_1,
                            profile.address_line_2,
                            profile.city,
                            profile.state_province,
                            profile.postal_code,
                            profile.country
                        ]
                        data[field] = ', '.join([part for part in address_parts if part])
                    elif field == 'city':
                        data[field] = profile.city or ''
                    elif field == 'country':
                        data[field] = profile.country or ''
                    elif field == 'emergency_contact_name':
                        data[field] = profile.emergency_contact_name or ''
                    elif field == 'emergency_contact_phone':
                        data[field] = profile.emergency_contact_phone or ''
                    elif field == 'education_level':
                        data[field] = profile.get_education_level_display() if profile.education_level else ''
                    elif field == 'employment_status':
                        data[field] = profile.get_employment_status_display() if profile.employment_status else ''
                    elif field == 'occupation':
                        data[field] = profile.occupation or ''
                    elif field == 'languages_spoken':
                        data[field] = ', '.join(profile.languages_spoken) if profile.languages_spoken else ''
                    elif field == 'nationality':
                        data[field] = profile.nationality or ''
                    else:
                        data[field] = ''
                
                # Recruitment preferences fields
                elif hasattr(submission, 'recruitment_preferences') and submission.recruitment_preferences:
                    recruitment = submission.recruitment_preferences
                    if field == 'volunteer_experience_level':
                        data[field] = recruitment.get_volunteer_experience_level_display() if recruitment.volunteer_experience_level else ''
                    elif field == 'previous_events':
                        data[field] = recruitment.previous_events or ''
                    elif field == 'special_skills':
                        data[field] = ', '.join(recruitment.special_skills) if recruitment.special_skills else ''
                    elif field == 'motivation':
                        data[field] = recruitment.motivation or ''
                    elif field == 'preferred_sports':
                        data[field] = ', '.join(recruitment.preferred_sports) if recruitment.preferred_sports else ''
                    elif field == 'preferred_venues':
                        data[field] = ', '.join(recruitment.preferred_venues) if recruitment.preferred_venues else ''
                    elif field == 'preferred_roles':
                        data[field] = ', '.join(recruitment.preferred_roles) if recruitment.preferred_roles else ''
                    elif field == 'availability_level':
                        data[field] = recruitment.get_availability_level_display() if recruitment.availability_level else ''
                    elif field == 'max_hours_per_day':
                        data[field] = str(recruitment.max_hours_per_day) if recruitment.max_hours_per_day else ''
                    elif field == 'has_own_transport':
                        data[field] = 'Yes' if recruitment.has_own_transport else 'No'
                    elif field == 'transport_method':
                        data[field] = recruitment.transport_method or ''
                    else:
                        data[field] = ''
                
                # Games information fields
                elif hasattr(submission, 'games_information') and submission.games_information:
                    games = submission.games_information
                    if field == 't_shirt_size':
                        data[field] = games.get_t_shirt_size_display() if games.t_shirt_size else ''
                    elif field == 'dietary_requirements':
                        data[field] = games.dietary_requirements or ''
                    elif field == 'has_food_allergies':
                        data[field] = 'Yes' if games.has_food_allergies else 'No'
                    elif field == 'food_allergy_details':
                        data[field] = games.food_allergy_details or ''
                    elif field == 'requires_accommodation':
                        data[field] = 'Yes' if games.requires_accommodation else 'No'
                    elif field == 'preferred_shifts':
                        data[field] = ', '.join(games.preferred_shifts) if games.preferred_shifts else ''
                    elif field == 'is_part_of_group':
                        data[field] = 'Yes' if games.is_part_of_group else 'No'
                    elif field == 'group_name':
                        data[field] = games.group_name or ''
                    elif field == 'how_did_you_hear':
                        data[field] = games.how_did_you_hear or ''
                    else:
                        data[field] = ''
                else:
                    data[field] = ''
                    
            except Exception as e:
                logger.warning(f"Error extracting field {field}: {e}")
                data[field] = ''
        
        return data
    
    def export_csv(self, queryset, fields):
        """Export data as CSV"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="eoi_export_{self.timestamp}.csv"'
        
        field_definitions = self.get_field_definitions()
        headers = [field_definitions[field]['label'] for field in fields]
        
        writer = csv.writer(response)
        writer.writerow(headers)
        
        for submission in queryset:
            data = self.extract_data(submission, fields)
            row = [data.get(field, '') for field in fields]
            writer.writerow(row)
        
        return response
    
    def export_excel(self, queryset, fields):
        """Export data as Excel file"""
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export")
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "EOI Submissions"
        
        field_definitions = self.get_field_definitions()
        headers = [field_definitions[field]['label'] for field in fields]
        
        # Style definitions
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Write data
        for row_num, submission in enumerate(queryset, 2):
            data = self.extract_data(submission, fields)
            for col, field in enumerate(fields, 1):
                cell = ws.cell(row=row_num, column=col, value=data.get(field, ''))
                cell.border = border
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 15
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="eoi_export_{self.timestamp}.xlsx"'
        
        return response
    
    def export_pdf(self, queryset, fields):
        """Export data as PDF"""
        if not PDF_AVAILABLE:
            raise ImportError("reportlab is required for PDF export")
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="eoi_export_{self.timestamp}.pdf"'
        
        # Create PDF document
        doc = SimpleDocTemplate(response, pagesize=A4)
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        
        # Title
        title = Paragraph(f"EOI Submissions Export - {datetime.now().strftime('%Y-%m-%d')}", title_style)
        elements.append(title)
        elements.append(Spacer(1, 12))
        
        # Prepare table data
        field_definitions = self.get_field_definitions()
        headers = [field_definitions[field]['label'] for field in fields]
        
        table_data = [headers]
        
        for submission in queryset:
            data = self.extract_data(submission, fields)
            row = [str(data.get(field, ''))[:50] + '...' if len(str(data.get(field, ''))) > 50 else str(data.get(field, '')) for field in fields]
            table_data.append(row)
        
        # Create table
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.green),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        
        return response
    
    def export_data(self, filters=None, fields=None, format_type=None):
        """Main export method"""
        if format_type:
            self.export_format = format_type.lower()
        
        # Get queryset
        queryset = self.get_base_queryset(filters)
        
        # Default fields if none specified
        if not fields:
            fields = [
                'submission_id', 'volunteer_type', 'status', 'created_at',
                'first_name', 'last_name', 'email', 'phone_number',
                'volunteer_experience_level', 'preferred_roles'
            ]
        
        # Log export activity
        log_audit_event(
            user=self.user,
            action='EOI_DATA_EXPORT',
            resource_type='EOISubmission',
            resource_id='bulk',
            details={
                'format': self.export_format,
                'record_count': queryset.count(),
                'fields': fields,
                'filters': filters or {}
            }
        )
        
        # Export based on format
        if self.export_format == 'csv':
            return self.export_csv(queryset, fields)
        elif self.export_format == 'excel':
            return self.export_excel(queryset, fields)
        elif self.export_format == 'pdf':
            return self.export_pdf(queryset, fields)
        else:
            raise ValueError(f"Unsupported export format: {self.export_format}")


def get_export_statistics(user):
    """Get statistics for export dashboard"""
    if not IsStaffOrVMTOrCVT().has_permission(type('MockRequest', (), {'user': user})(), None):
        raise PermissionDenied("User does not have permission to view statistics")
    
    stats = {
        'total_submissions': EOISubmission.objects.count(),
        'by_status': dict(EOISubmission.objects.values('status').annotate(count=Count('id')).values_list('status', 'count')),
        'by_volunteer_type': dict(EOISubmission.objects.values('volunteer_type').annotate(count=Count('id')).values_list('volunteer_type', 'count')),
        'recent_submissions': EOISubmission.objects.filter(
            created_at__gte=datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        ).count(),
        'completion_rate': EOISubmission.objects.filter(
            profile_section_complete=True,
            recruitment_section_complete=True,
            games_section_complete=True
        ).count() / max(EOISubmission.objects.count(), 1) * 100
    }
    
    return stats 