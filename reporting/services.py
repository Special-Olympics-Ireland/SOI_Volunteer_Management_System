"""
Report generation services with comprehensive export capabilities.
Supports CSV, PDF, Excel, and JSON formats for various report types.
"""

import csv
import json
import io
import os
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional, Union
from django.conf import settings
from django.utils import timezone
from django.db.models import QuerySet, Count, Q, Avg, Sum
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.contrib.auth import get_user_model

# Import for Excel support
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Import for PDF support
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

from .models import Report, ReportMetrics
from .base import BaseReportGenerator
from .report_generators import get_report_generator_class
from volunteers.models import VolunteerProfile
from events.models import Event, Venue, Role
from accounts.models import User

User = get_user_model()


class ReportGenerationError(Exception):
    """Custom exception for report generation errors"""
    pass





class CSVReportGenerator(BaseReportGenerator):
    """
    CSV report generator with advanced formatting and filtering.
    """
    
    def __init__(self, report: Report):
        super().__init__(report)
        # Get the specific report generator for data access
        self.specific_generator = None
        generator_class = get_report_generator_class(self.report.report_type)
        if generator_class:
            self.specific_generator = generator_class(self.report)
    
    def get_queryset(self):
        """Get queryset from specific generator"""
        if self.specific_generator:
            return self.specific_generator.get_queryset()
        # Fallback to empty queryset
        from django.contrib.auth.models import User
        return User.objects.none()
    
    def get_columns(self):
        """Get columns from specific generator"""
        if self.specific_generator:
            return self.specific_generator.get_columns()
        # Fallback columns
        return [{'key': 'id', 'label': 'ID', 'type': 'text'}]
    
    def format_row_data(self, obj):
        """Format row data using specific generator"""
        if self.specific_generator:
            return self.specific_generator.format_row_data(obj)
        # Fallback formatting
        return [str(obj.id) if hasattr(obj, 'id') else str(obj)]
    
    def generate(self) -> str:
        """Generate CSV report and return file path"""
        try:
            self.start_generation()
            self.update_progress(10, "Preparing data...")
            
            # Get data
            queryset = self.get_queryset()
            columns = self.get_columns()
            
            self.update_progress(30, "Processing records...")
            
            # Create file path
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{self.report.report_type.lower()}_{timestamp}.csv"
            file_path = os.path.join(settings.MEDIA_ROOT, 'reports', filename)
            
            # Ensure directory exists
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            
            # Generate CSV
            with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                
                # Write header
                headers = [col['label'] for col in columns]
                writer.writerow(headers)
                
                # Write data rows
                total_records = queryset.count()
                processed = 0
                
                for obj in queryset.iterator():
                    row_data = self.format_row_data(obj)
                    writer.writerow(row_data)
                    
                    processed += 1
                    if processed % 100 == 0:  # Update progress every 100 records
                        progress = 30 + int((processed / total_records) * 60)
                        self.update_progress(progress, f"Processed {processed}/{total_records} records")
            
            # Get file size
            file_size = os.path.getsize(file_path)
            
            self.update_progress(95, "Finalizing report...")
            self.complete_generation(file_path, file_size, total_records)
            
            return file_path
            
        except Exception as e:
            self.fail_generation(f"CSV generation failed: {str(e)}")
            raise ReportGenerationError(f"Failed to generate CSV report: {str(e)}")


class ExcelReportGenerator(BaseReportGenerator):
    """
    Excel report generator with advanced formatting and styling.
    """
    
    def generate(self) -> str:
        """Generate Excel report and return file path"""
        if not EXCEL_AVAILABLE:
            raise ReportGenerationError("Excel support not available. Install openpyxl.")
        
        try:
            self.start_generation()
            self.update_progress(10, "Preparing Excel workbook...")
            
            # Get data
            queryset = self.get_queryset()
            columns = self.get_columns()
            
            self.update_progress(30, "Creating Excel structure...")
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = self.report.name[:31]  # Excel sheet name limit
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")  # SOI Green
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Write headers
            headers = [col['label'] for col in columns]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            self.update_progress(40, "Writing data rows...")
            
            # Write data rows
            total_records = queryset.count()
            processed = 0
            row_idx = 2
            
            for obj in queryset.iterator():
                row_data = self.format_row_data(obj)
                
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    
                    # Apply column-specific formatting
                    col_info = columns[col_idx - 1]
                    if col_info.get('type') == 'number':
                        cell.number_format = '#,##0'
                    elif col_info.get('type') == 'currency':
                        cell.number_format = '$#,##0.00'
                    elif col_info.get('type') == 'percentage':
                        cell.number_format = '0.00%'
                    elif col_info.get('type') == 'date':
                        cell.number_format = 'yyyy-mm-dd'
                    elif col_info.get('type') == 'datetime':
                        cell.number_format = 'yyyy-mm-dd hh:mm:ss'
                
                row_idx += 1
                processed += 1
                
                if processed % 100 == 0:  # Update progress every 100 records
                    progress = 40 + int((processed / total_records) * 50)
                    self.update_progress(progress, f"Processed {processed}/{total_records} records")
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add metadata sheet
            meta_ws = wb.create_sheet("Report Metadata")
            metadata = [
                ["Report Name", self.report.name],
                ["Report Type", self.report.get_report_type_display()],
                ["Generated By", self.report.created_by.get_full_name()],
                ["Generated At", timezone.now().strftime("%Y-%m-%d %H:%M:%S")],
                ["Total Records", total_records],
                ["Parameters", json.dumps(self.parameters, indent=2)]
            ]
            
            for row_idx, (key, value) in enumerate(metadata, 1):
                meta_ws.cell(row=row_idx, column=1, value=key).font = Font(bold=True)
                meta_ws.cell(row=row_idx, column=2, value=str(value))
            
            # Create file path
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{self.report.report_type.lower()}_{timestamp}.xlsx"
            file_path = os.path.join(settings.MEDIA_ROOT, 'reports', filename)
            
            # Ensure directory exists
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            
            # Save workbook
            self.update_progress(95, "Saving Excel file...")
            wb.save(file_path)
            
            # Get file size
            file_size = os.path.getsize(file_path)
            
            self.complete_generation(file_path, file_size, total_records)
            
            return file_path
            
        except Exception as e:
            self.fail_generation(f"Excel generation failed: {str(e)}")
            raise ReportGenerationError(f"Failed to generate Excel report: {str(e)}")


class PDFReportGenerator(BaseReportGenerator):
    """
    PDF report generator with professional formatting.
    """
    
    def generate(self) -> str:
        """Generate PDF report and return file path"""
        if not PDF_AVAILABLE:
            raise ReportGenerationError("PDF support not available. Install reportlab.")
        
        try:
            self.start_generation()
            self.update_progress(10, "Preparing PDF document...")
            
            # Get data
            queryset = self.get_queryset()
            columns = self.get_columns()
            
            # Create file path
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{self.report.report_type.lower()}_{timestamp}.pdf"
            file_path = os.path.join(settings.MEDIA_ROOT, 'reports', filename)
            
            # Ensure directory exists
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            
            # Create PDF document
            doc = SimpleDocTemplate(file_path, pagesize=A4)
            story = []
            
            # Get styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=30,
                textColor=colors.Color(34/255, 139/255, 34/255)  # SOI Green
            )
            
            # Add title
            title = Paragraph(f"{self.report.name}", title_style)
            story.append(title)
            
            # Add metadata
            metadata_text = f"""
            <b>Report Type:</b> {self.report.get_report_type_display()}<br/>
            <b>Generated By:</b> {self.report.created_by.get_full_name()}<br/>
            <b>Generated At:</b> {timezone.now().strftime("%Y-%m-%d %H:%M:%S")}<br/>
            <b>Total Records:</b> {queryset.count()}
            """
            metadata = Paragraph(metadata_text, styles['Normal'])
            story.append(metadata)
            story.append(Spacer(1, 20))
            
            self.update_progress(30, "Building PDF table...")
            
            # Prepare table data
            headers = [col['label'] for col in columns]
            table_data = [headers]
            
            total_records = queryset.count()
            processed = 0
            
            # Add data rows (limit for PDF to prevent huge files)
            max_rows = min(total_records, 1000)  # Limit PDF to 1000 rows
            
            for obj in queryset[:max_rows]:
                row_data = self.format_row_data(obj)
                # Truncate long text for PDF display
                formatted_row = []
                for item in row_data:
                    if isinstance(item, str) and len(item) > 50:
                        formatted_row.append(item[:47] + "...")
                    else:
                        formatted_row.append(str(item) if item is not None else "")
                
                table_data.append(formatted_row)
                processed += 1
                
                if processed % 50 == 0:  # Update progress every 50 records
                    progress = 30 + int((processed / max_rows) * 60)
                    self.update_progress(progress, f"Processed {processed}/{max_rows} records")
            
            # Create table
            table = Table(table_data)
            
            # Style the table
            table_style = TableStyle([
                # Header styling
                ('BACKGROUND', (0, 0), (-1, 0), colors.Color(34/255, 139/255, 34/255)),  # SOI Green
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                
                # Data styling
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                
                # Grid
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                
                # Alternating row colors
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ])
            
            table.setStyle(table_style)
            story.append(table)
            
            # Add note if data was truncated
            if total_records > max_rows:
                note_text = f"<i>Note: This PDF shows the first {max_rows} records out of {total_records} total records. For complete data, please use CSV or Excel export.</i>"
                note = Paragraph(note_text, styles['Italic'])
                story.append(Spacer(1, 20))
                story.append(note)
            
            self.update_progress(90, "Building PDF document...")
            
            # Build PDF
            doc.build(story)
            
            # Get file size
            file_size = os.path.getsize(file_path)
            
            self.complete_generation(file_path, file_size, min(total_records, max_rows))
            
            return file_path
            
        except Exception as e:
            self.fail_generation(f"PDF generation failed: {str(e)}")
            raise ReportGenerationError(f"Failed to generate PDF report: {str(e)}")


class JSONReportGenerator(BaseReportGenerator):
    """
    JSON report generator for API consumption and data exchange.
    """
    
    def generate(self) -> str:
        """Generate JSON report and return file path"""
        try:
            self.start_generation()
            self.update_progress(10, "Preparing JSON structure...")
            
            # Get data
            queryset = self.get_queryset()
            columns = self.get_columns()
            
            self.update_progress(30, "Serializing data...")
            
            # Prepare JSON structure
            report_data = {
                'metadata': {
                    'report_name': self.report.name,
                    'report_type': self.report.report_type,
                    'generated_by': self.report.created_by.username,
                    'generated_at': timezone.now().isoformat(),
                    'parameters': self.parameters,
                    'total_records': queryset.count()
                },
                'columns': columns,
                'data': []
            }
            
            # Process data
            total_records = queryset.count()
            processed = 0
            
            for obj in queryset.iterator():
                row_data = self.format_row_data(obj)
                
                # Create row dictionary
                row_dict = {}
                for idx, col in enumerate(columns):
                    value = row_data[idx] if idx < len(row_data) else None
                    
                    # Handle datetime serialization
                    if hasattr(value, 'isoformat'):
                        value = value.isoformat()
                    elif value is None:
                        value = None
                    else:
                        value = str(value)
                    
                    row_dict[col['key']] = value
                
                report_data['data'].append(row_dict)
                
                processed += 1
                if processed % 100 == 0:  # Update progress every 100 records
                    progress = 30 + int((processed / total_records) * 60)
                    self.update_progress(progress, f"Processed {processed}/{total_records} records")
            
            # Create file path
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{self.report.report_type.lower()}_{timestamp}.json"
            file_path = os.path.join(settings.MEDIA_ROOT, 'reports', filename)
            
            # Ensure directory exists
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            
            # Write JSON file
            self.update_progress(90, "Writing JSON file...")
            with open(file_path, 'w', encoding='utf-8') as jsonfile:
                json.dump(report_data, jsonfile, indent=2, ensure_ascii=False)
            
            # Get file size
            file_size = os.path.getsize(file_path)
            
            self.complete_generation(file_path, file_size, total_records)
            
            return file_path
            
        except Exception as e:
            self.fail_generation(f"JSON generation failed: {str(e)}")
            raise ReportGenerationError(f"Failed to generate JSON report: {str(e)}")


class ReportGeneratorFactory:
    """
    Factory class to create appropriate report generators based on export format.
    """
    
    GENERATORS = {
        Report.ExportFormat.CSV: CSVReportGenerator,
        Report.ExportFormat.EXCEL: ExcelReportGenerator,
        Report.ExportFormat.PDF: PDFReportGenerator,
        Report.ExportFormat.JSON: JSONReportGenerator,
    }
    
    @classmethod
    def create_generator(cls, report: Report) -> BaseReportGenerator:
        """Create appropriate generator for the report format"""
        generator_class = cls.GENERATORS.get(report.export_format)
        
        if not generator_class:
            raise ReportGenerationError(f"Unsupported export format: {report.export_format}")
        
        # Check format-specific requirements
        if report.export_format == Report.ExportFormat.EXCEL and not EXCEL_AVAILABLE:
            raise ReportGenerationError("Excel export requires openpyxl package")
        
        if report.export_format == Report.ExportFormat.PDF and not PDF_AVAILABLE:
            raise ReportGenerationError("PDF export requires reportlab package")
        
        return generator_class(report)


def generate_report(report_id: str) -> str:
    """
    Main function to generate a report by ID.
    Returns the file path of the generated report.
    """
    try:
        report = Report.objects.get(id=report_id)
        
        # Create appropriate generator
        generator = ReportGeneratorFactory.create_generator(report)
        
        # Generate the report
        file_path = generator.generate()
        
        return file_path
        
    except Report.DoesNotExist:
        raise ReportGenerationError(f"Report with ID {report_id} not found")
    except Exception as e:
        raise ReportGenerationError(f"Report generation failed: {str(e)}")


def cleanup_expired_reports():
    """
    Cleanup expired reports and their files.
    Should be run periodically via cron job or task queue.
    """
    expired_reports = Report.objects.filter(
        expires_at__lt=timezone.now(),
        status=Report.Status.COMPLETED
    )
    
    cleaned_count = 0
    for report in expired_reports:
        try:
            # Delete file if it exists
            if report.file_path and os.path.exists(report.file_path):
                os.remove(report.file_path)
            
            # Update report status
            report.status = Report.Status.EXPIRED
            report.file_path = ""
            report.file_size = 0
            report.save()
            
            cleaned_count += 1
            
        except Exception as e:
            # Log error but continue with other reports
            print(f"Error cleaning up report {report.id}: {str(e)}")
    
    return cleaned_count 